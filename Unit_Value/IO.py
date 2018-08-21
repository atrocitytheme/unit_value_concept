import re
import xlrd
import codecs
import openpyxl

def build_excel(file_path, data_list, field_title=[], hyper_index_set=set()):
    work_book = openpyxl.Workbook()
    work_book.create_sheet(u'sheet1')
    work_book.remove(work_book.get_sheet_by_name('Sheet'))
    column_buffer = 0
    if field_title:
        for f_index, f_title in enumerate(field_title):
            work_book.get_sheet_by_name(u'sheet1').cell(row=1, column=f_index + 1, value=f_title)
        column_buffer = 1

    for d_index, data_item in enumerate(data_list):
        for df_index, d_element in enumerate(data_item):
            df_index = df_index + 1
            cell_value = work_book.get_sheet_by_name(u'sheet1').cell(row=d_index + 1 + column_buffer, column=df_index , value=d_element)
            if hyper_index_set and df_index in hyper_index_set:
                cell_value.hyperlink = d_element

    work_book.save(file_path)

def load_data_from_excel_with_sheet(excel_file, sheet_index = [], sheet_name=[], column_indexes=[], filter_header=False):
    if not excel_file or (not sheet_index and not sheet_name):
        raise Exception('please set the parameter: excel_file column_indexes and \
                         one of the sheet_index and the sheet_name to  load data')
    xlrd_file = xlrd.open_workbook(excel_file)
    if sheet_index:
        sheet_data = [(s_i, xlrd_file.sheet_by_index(s_i)) for s_i in sheet_index]
    elif sheet_name:
        sheet_data = [(s_n, xlrd_file.sheet_by_name(s_n)) for s_n in sheet_name]

    sheet_ds = dict()
    for sheet,data in sheet_data:
        result_data = list()
        if filter_header:
            row_range = range(data.nrows)[1:]
        else:
            row_range = range(data.nrows)
        col_indexes = range(data.ncols)
        real_cols = set(col_indexes)
        for n_row in row_range:
            temp_list = []
            if not column_indexes:
                column_indexes = col_indexes
            for c_i in column_indexes:
                if c_i in real_cols:
                    temp_list.append(data.cell(n_row, c_i).value)
                else:
                    temp_list.append(u'')

            result_data.append(temp_list)
        sheet_ds[sheet] = result_data
    return sheet_ds

def excel2txt(excel_file, out_txt, sep='\t'):
    data_l = load_data_from_excel_with_sheet(excel_file, sheet_index=[0])[0]
    with codecs.open(out_txt, 'w', encoding='utf-8') as fw:
        for l in data_l:
            fw.write('{}\n'.format(sep.join([x.replace(sep, ' ') for x in l])))

def excel2txt_win(excel_file, out_win_txt, sep='\t'):
    data_l = load_data_from_excel_with_sheet(excel_file, sheet_index=[0])[0]
    with codecs.open(out_win_txt, 'w', encoding='utf-8') as fw:
        fw.write(codecs.BOM_UTF8.decode('utf-8'))
        for l in data_l:
            fw.write('{}\n'.format(sep.join([x.replace(sep, ' ') for x in l])))

def dataLst2txt(data_l, out_txt, sep='\t', replace_sep=' '):
    with codecs.open(out_txt, 'w', encoding='utf-8') as fw:
        for l in data_l:
            write_con = l if isinstance(l, str) else sep.join([str(x).replace(sep, replace_sep) for x in l])
            fw.write('{}\n'.format(write_con))

def readtxt2Lst(data_file, sep='\t'):
    with codecs.open(data_file, 'r', encoding='utf-8') as fr:
        data_l = [l.split(sep)[0] if len(l.split(sep)) == 1 else l.split(sep) for l in fr.read().splitlines()]
    return data_l

def txt2excel(data_file, out_excel, sep='\t'):
    data_l = readtxt2Lst(data_file,sep)
    build_excel(out_excel, data_l)
